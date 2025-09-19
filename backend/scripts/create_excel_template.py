#!/usr/bin/env python3
"""
Script para crear un archivo Excel de ejemplo con el plan de cuentas
"""

import pandas as pd
import os
from datetime import datetime

# Datos de ejemplo para el plan de cuentas
SAMPLE_DATA = [
    {"Código": "1101", "Cuenta": "CAJA", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 5000.00, "Saldo Crédito": 0.00},
    {"Código": "1102", "Cuenta": "BANCOS", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 25000.00, "Saldo Crédito": 0.00},
    {"Código": "1103", "Cuenta": "INVERSIONES TEMPORALES", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 10000.00, "Saldo Crédito": 0.00},
    {"Código": "1201", "Cuenta": "CUENTAS POR COBRAR COMERCIALES", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 15000.00, "Saldo Crédito": 0.00},
    {"Código": "1202", "Cuenta": "CUENTAS POR COBRAR NO COMERCIALES", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 5000.00, "Saldo Crédito": 0.00},
    {"Código": "1301", "Cuenta": "MERCADERÍAS", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 30000.00, "Saldo Crédito": 0.00},
    {"Código": "1302", "Cuenta": "MATERIA PRIMA", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 20000.00, "Saldo Crédito": 0.00},
    {"Código": "1401", "Cuenta": "INMUEBLES, PLANTA Y EQUIPO", "Tipo": "activo", "Naturaleza": "deudora", "Saldo Débito": 100000.00, "Saldo Crédito": 0.00},
    {"Código": "1402", "Cuenta": "DEPRECIACIÓN ACUMULADA", "Tipo": "activo", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 20000.00},
    {"Código": "2101", "Cuenta": "CUENTAS POR PAGAR COMERCIALES", "Tipo": "pasivo", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 12000.00},
    {"Código": "2102", "Cuenta": "CUENTAS POR PAGAR NO COMERCIALES", "Tipo": "pasivo", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 8000.00},
    {"Código": "2103", "Cuenta": "IMPUESTOS POR PAGAR", "Tipo": "pasivo", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 5000.00},
    {"Código": "2201", "Cuenta": "OBLIGACIONES FINANCIERAS", "Tipo": "pasivo", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 50000.00},
    {"Código": "3101", "Cuenta": "CAPITAL SOCIAL", "Tipo": "patrimonio", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 100000.00},
    {"Código": "3102", "Cuenta": "RESERVAS", "Tipo": "patrimonio", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 15000.00},
    {"Código": "3201", "Cuenta": "UTILIDADES ACUMULADAS", "Tipo": "patrimonio", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 25000.00},
    {"Código": "4101", "Cuenta": "VENTAS", "Tipo": "ingresos", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "4102", "Cuenta": "SERVICIOS", "Tipo": "ingresos", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "4201", "Cuenta": "INGRESOS FINANCIEROS", "Tipo": "ingresos", "Naturaleza": "acreedora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "5101", "Cuenta": "GASTOS DE VENTAS", "Tipo": "gastos", "Naturaleza": "deudora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "5102", "Cuenta": "GASTOS ADMINISTRATIVOS", "Tipo": "gastos", "Naturaleza": "deudora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "5201", "Cuenta": "GASTOS FINANCIEROS", "Tipo": "gastos", "Naturaleza": "deudora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "6101", "Cuenta": "COSTO DE MERCADERÍAS", "Tipo": "costos", "Naturaleza": "deudora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
    {"Código": "6201", "Cuenta": "MATERIA PRIMA CONSUMIDA", "Tipo": "costos", "Naturaleza": "deudora", "Saldo Débito": 0.00, "Saldo Crédito": 0.00},
]

def create_excel_template():
    """Crear archivo Excel de plantilla"""
    
    # Crear DataFrame
    df = pd.DataFrame(SAMPLE_DATA)
    
    # Crear archivo Excel con formato
    filename = f"plan_cuentas_ejemplo_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Escribir datos principales
        df.to_excel(writer, sheet_name='Plan de Cuentas', index=False)
        
        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Plan de Cuentas']
        
        # Formatear encabezados
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Aplicar formato a los encabezados
        for col in range(1, 7):  # Columnas A-F
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12,  # Código
            'B': 35,  # Cuenta
            'C': 15,  # Tipo
            'D': 15,  # Naturaleza
            'E': 15,  # Saldo Débito
            'F': 15   # Saldo Crédito
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatear números como moneda
        from openpyxl.styles import NamedStyle
        
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '"$"#,##0.00'
        
        # Aplicar formato de moneda a las columnas de saldos
        for row in range(2, len(SAMPLE_DATA) + 2):
            worksheet.cell(row=row, column=5).number_format = '"$"#,##0.00'  # Saldo Débito
            worksheet.cell(row=row, column=6).number_format = '"$"#,##0.00'  # Saldo Crédito
        
        # Agregar bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(SAMPLE_DATA) + 2):
            for col in range(1, 7):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Crear hoja de instrucciones
        instructions_sheet = workbook.create_sheet("Instrucciones")
        
        instructions = [
            "INSTRUCCIONES PARA USAR ESTA PLANTILLA",
            "",
            "1. Complete los saldos iniciales en las columnas 'Saldo Débito' y 'Saldo Crédito'",
            "2. Los tipos de cuenta válidos son: activo, pasivo, patrimonio, ingresos, gastos, costos",
            "3. Las naturalezas válidas son: deudora, acreedora",
            "4. Solo complete los saldos en las cuentas que tengan movimientos iniciales",
            "5. El total de débitos debe ser igual al total de créditos",
            "6. Guarde el archivo y úselo para importar en el sistema",
            "",
            "EJEMPLO DE CUENTAS CON SALDOS INICIALES:",
            "Código | Cuenta | Tipo | Naturaleza | Saldo Débito | Saldo Crédito",
            "1101 | CAJA | activo | deudora | 5000.00 | 0.00",
            "1102 | BANCOS | activo | deudora | 25000.00 | 0.00",
            "2101 | CUENTAS POR PAGAR | pasivo | acreedora | 0.00 | 12000.00",
            "3101 | CAPITAL SOCIAL | patrimonio | acreedora | 0.00 | 100000.00",
            "",
            "NOTAS IMPORTANTES:",
            "- No modifique los códigos de cuenta",
            "- No modifique los nombres de las cuentas",
            "- Solo modifique los saldos iniciales",
            "- Los saldos deben estar en formato numérico (ej: 1000.50)",
            "- Use punto como separador decimal"
        ]
        
        for i, instruction in enumerate(instructions, 1):
            instructions_sheet.cell(row=i, column=1, value=instruction)
        
        # Formatear hoja de instrucciones
        instructions_sheet.column_dimensions['A'].width = 80
        
        # Formatear título
        title_cell = instructions_sheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        
        # Formatear encabezados de ejemplo
        for col in range(1, 7):
            cell = instructions_sheet.cell(row=10, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    print(f"✅ Archivo Excel creado: {filename}")
    print(f"📁 Ubicación: {os.path.abspath(filename)}")
    print(f"📊 Total de cuentas: {len(SAMPLE_DATA)}")
    
    # Calcular totales
    total_debit = sum(row["Saldo Débito"] for row in SAMPLE_DATA)
    total_credit = sum(row["Saldo Crédito"] for row in SAMPLE_DATA)
    
    print(f"\n💰 Resumen de saldos:")
    print(f"   Total Débito: ${total_debit:,.2f}")
    print(f"   Total Crédito: ${total_credit:,.2f}")
    print(f"   Diferencia: ${abs(total_debit - total_credit):,.2f}")
    
    if abs(total_debit - total_credit) < 0.01:
        print("✅ El plan de cuentas está balanceado")
    else:
        print("⚠️  El plan de cuentas no está balanceado")

if __name__ == "__main__":
    create_excel_template()
